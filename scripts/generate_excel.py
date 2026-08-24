#!/usr/bin/env python3
"""
AWS Inventory Excel Report Generator

Reads inventory-data.json and produces a formatted Excel workbook
with one sheet per service category, a Summary sheet, and a ScanNotes sheet.

Usage:
    python scripts/generate_excel.py [--input PATH] [--output PATH]

Dependencies:
    pip install openpyxl
"""

import json
import sys
import os
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


# ─── Styling Constants ────────────────────────────────────────────────────────

HEADER_FILL_COLOR = "1F4E79"
HEADER_FONT_COLOR = "FFFFFF"
ALT_ROW_FILL_COLOR = "F2F7FC"
TITLE_FONT_SIZE = 16
SUBTITLE_FONT_SIZE = 14
BODY_FONT = "Calibri"
BODY_FONT_SIZE = 11
MIN_COL_WIDTH = 12
MAX_COL_WIDTH = 50
MAX_CELL_LENGTH = 32767  # Excel cell character limit


# ─── Styles ───────────────────────────────────────────────────────────────────

header_font = Font(name=BODY_FONT, size=BODY_FONT_SIZE, bold=True, color=HEADER_FONT_COLOR)
header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
alt_row_fill = PatternFill(start_color=ALT_ROW_FILL_COLOR, end_color=ALT_ROW_FILL_COLOR, fill_type="solid")
title_font = Font(name=BODY_FONT, size=TITLE_FONT_SIZE, bold=True)
subtitle_font = Font(name=BODY_FONT, size=SUBTITLE_FONT_SIZE, bold=True)
body_font = Font(name=BODY_FONT, size=BODY_FONT_SIZE)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def sanitize_sheet_name(name: str) -> str:
    """Ensure sheet name is valid for Excel (max 31 chars, no special chars)."""
    invalid_chars = ["[", "]", ":", "*", "?", "/", "\\"]
    for ch in invalid_chars:
        name = name.replace(ch, "_")
    return name[:31]


def truncate_cell_value(value) -> str:
    """Truncate cell value if it exceeds Excel's character limit."""
    if value is None:
        return ""
    s = str(value)
    if len(s) > MAX_CELL_LENGTH:
        return s[: MAX_CELL_LENGTH - 3] + "..."
    return s


def auto_fit_column_width(ws, col_idx: int, max_rows: int = 100):
    """Calculate optimal column width based on content (sample first N rows)."""
    max_length = MIN_COL_WIDTH
    col_letter = get_column_letter(col_idx)

    for row_idx, row in enumerate(ws.iter_rows(min_col=col_idx, max_col=col_idx), 1):
        if row_idx > max_rows:
            break
        cell = row[0]
        if cell.value:
            cell_length = len(str(cell.value))
            max_length = max(max_length, min(cell_length + 2, MAX_COL_WIDTH))

    ws.column_dimensions[col_letter].width = max_length


def human_readable_size(size_bytes) -> str:
    """Convert bytes to human-readable format."""
    if not size_bytes or size_bytes == "":
        return ""
    try:
        size_bytes = int(size_bytes)
    except (ValueError, TypeError):
        return str(size_bytes)

    if size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 ** 2:
        return f"{size_bytes / 1024:.1f} KB"
    elif size_bytes < 1024 ** 3:
        return f"{size_bytes / (1024 ** 2):.1f} MB"
    elif size_bytes < 1024 ** 4:
        return f"{size_bytes / (1024 ** 3):.2f} GB"
    else:
        return f"{size_bytes / (1024 ** 4):.2f} TB"


def create_summary_sheet(wb: Workbook, metadata: dict, categories: dict):
    """Create the Summary sheet with scan overview and resource counts."""
    ws = wb.active
    ws.title = "Summary"

    # Title
    ws.merge_cells("A1:E1")
    ws["A1"] = "AWS Infrastructure Inventory Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    # Metadata
    ws["A3"] = "Account ID:"
    ws["B3"] = metadata.get("accountId", "N/A")
    ws["A3"].font = Font(bold=True)

    ws["A4"] = "Caller Identity:"
    ws["B4"] = metadata.get("callerArn", "N/A")
    ws["A4"].font = Font(bold=True)

    ws["A5"] = "Scan Date:"
    ws["B5"] = metadata.get("scanDate", "N/A")
    ws["A5"].font = Font(bold=True)

    ws["A6"] = "Regions Scanned:"
    ws["B6"] = ", ".join(metadata.get("regions", []))
    ws["A6"].font = Font(bold=True)

    ws["A7"] = "Total Resources:"
    ws["B7"] = metadata.get("totalResources", 0)
    ws["A7"].font = Font(bold=True)

    # Resource summary table
    ws["A9"] = "Resource Summary by Category"
    ws["A9"].font = subtitle_font

    # Table headers
    headers = ["Category", "Resource Count", "Regions Found"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Table data — sorted by resource count descending
    category_counts = []
    for key, cat_data in categories.items():
        count = len(cat_data.get("data", []))
        display_name = cat_data.get("displayName", key)
        # Extract unique regions from data (first column is often Region)
        regions_set = set()
        for row in cat_data.get("data", []):
            if row and row[0] and row[0] not in ("Global", ""):
                regions_set.add(row[0])
        regions_str = ", ".join(sorted(regions_set)) if regions_set else "Global"
        category_counts.append((display_name, count, regions_str))

    category_counts.sort(key=lambda x: x[1], reverse=True)

    for row_idx, (name, count, regions) in enumerate(category_counts, 11):
        ws.cell(row=row_idx, column=1, value=name).border = thin_border
        ws.cell(row=row_idx, column=2, value=count).border = thin_border
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=regions).border = thin_border

        # Alternate row shading
        if row_idx % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).fill = alt_row_fill

    # Auto-fit columns
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15


def create_data_sheet(wb: Workbook, sheet_name: str, columns: list, data: list):
    """Create a formatted data sheet for a service category."""
    safe_name = sanitize_sheet_name(sheet_name)
    ws = wb.create_sheet(title=safe_name)

    # Write header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Write data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell_value = truncate_cell_value(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

            # Alternate row shading
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if data:
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}{len(data) + 1}"

    # Auto-fit column widths (sample first 100 rows for performance)
    row_count = len(data)
    if row_count <= 10000:
        for col_idx in range(1, len(columns) + 1):
            auto_fit_column_width(ws, col_idx, min(row_count + 1, 100))
    else:
        # Fixed widths for large datasets
        for col_idx in range(1, len(columns) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20


def create_scan_notes_sheet(wb: Workbook, scan_notes: list):
    """Create the ScanNotes sheet with errors and operational notes."""
    ws = wb.create_sheet(title="ScanNotes")

    # Headers
    headers = ["Timestamp", "Region", "Service", "Issue Type", "Details"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data
    for row_idx, note in enumerate(scan_notes, 2):
        ws.cell(row=row_idx, column=1, value=note.get("timestamp", "")).border = thin_border
        ws.cell(row=row_idx, column=2, value=note.get("region", "")).border = thin_border
        ws.cell(row=row_idx, column=3, value=note.get("service", "")).border = thin_border
        ws.cell(row=row_idx, column=4, value=note.get("issueType", "")).border = thin_border
        ws.cell(row=row_idx, column=5, value=note.get("details", "")).border = thin_border

        if row_idx % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).fill = alt_row_fill

    # Freeze and filter
    ws.freeze_panes = "A2"
    if scan_notes:
        ws.auto_filter.ref = f"A1:E{len(scan_notes) + 1}"

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 60


def generate_workbook(input_path: str, output_path: str = None) -> str:
    """
    Main function: reads inventory JSON and generates Excel workbook.

    Args:
        input_path: Path to inventory-data.json
        output_path: Optional explicit output path. If None, auto-generates filename.

    Returns:
        The path to the generated Excel file.
    """
    # Read input data
    with open(input_path, "r", encoding="utf-8") as f:
        inventory_data = json.load(f)

    metadata = inventory_data.get("metadata", {})
    categories = inventory_data.get("categories", {})
    scan_notes = inventory_data.get("scanNotes", [])

    # Determine output path
    if not output_path:
        account_id = metadata.get("accountId", "unknown")
        scan_date = metadata.get("scanDate", datetime.utcnow().isoformat())
        try:
            dt = datetime.fromisoformat(scan_date.replace("Z", "+00:00"))
            date_str = dt.strftime("%Y%m%d-%H%M")
        except (ValueError, AttributeError):
            date_str = datetime.utcnow().strftime("%Y%m%d-%H%M")

        output_dir = os.path.dirname(input_path)
        output_path = os.path.join(output_dir, f"aws-inventory-{account_id}-{date_str}.xlsx")

    # Ensure output directory exists
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    # Create workbook
    wb = Workbook()

    # 1. Summary sheet
    create_summary_sheet(wb, metadata, categories)

    # 2. Service category sheets (skip empty ones)
    for sheet_key, cat_data in categories.items():
        data = cat_data.get("data", [])
        if not data:
            continue  # Skip empty categories

        columns = cat_data.get("columns", [])
        if not columns:
            continue

        # Validate row lengths match column count
        validated_data = []
        col_count = len(columns)
        for row in data:
            if len(row) < col_count:
                row = row + [""] * (col_count - len(row))
            elif len(row) > col_count:
                row = row[:col_count]
            validated_data.append(row)

        create_data_sheet(wb, sheet_key, columns, validated_data)

    # 3. ScanNotes sheet
    create_scan_notes_sheet(wb, scan_notes)

    # Save workbook
    wb.save(output_path)
    print(f"Excel report generated: {output_path}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"  Total resources: {metadata.get('totalResources', 'N/A')}")
    print(f"  Categories with data: {sum(1 for c in categories.values() if c.get('data'))}")

    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Generate AWS Inventory Excel Report from JSON data"
    )
    parser.add_argument(
        "--input",
        "-i",
        default="inventory-reports/inventory-data.json",
        help="Path to inventory-data.json (default: inventory-reports/inventory-data.json)",
    )
    parser.add_argument(
        "--output",
        "-o",
        default=None,
        help="Output Excel file path (default: auto-generated in same directory as input)",
    )

    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: Input file not found: {args.input}")
        print("Run the inventory scan first to generate inventory-data.json")
        sys.exit(1)

    output_file = generate_workbook(args.input, args.output)
    print(f"\nDone! Open the report: {output_file}")


if __name__ == "__main__":
    main()
